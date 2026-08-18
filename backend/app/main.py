from __future__ import annotations

import hashlib
import hmac
import base64
import json
from io import BytesIO
import os
import re
import secrets
import time
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Annotated

import fitz
import jwt
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import Boolean, Date, DateTime, ForeignKey, Integer, String, Text, create_engine, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, sessionmaker

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+psycopg2://altura:altura@postgres:5432/altura_nexo",
)
JWT_SECRET = os.getenv("JWT_SECRET", "dev-only-secret")
DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
UPLOAD_DIR = DATA_DIR / "uploads"
REPORT_DIR = DATA_DIR / "reports"
OPENCODE_API_KEY = os.getenv("OPENCODE_API_KEY", "").strip()
OPENCODE_URL = os.getenv("OPENCODE_URL", "https://opencode.ai/zen/go/v1/responses").strip()
OPENCODE_MODEL = os.getenv("OPENCODE_MODEL", "gpt-5.6-luna").strip()
MONTHS = {
    1: "ENE",
    2: "FEB",
    3: "MAR",
    4: "ABR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AGO",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DIC",
}
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".pdf", ".jpg", ".jpeg", ".png"}

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)


class Base(DeclarativeBase):
    pass


class User(Base):
    __tablename__ = "users"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    username: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    role: Mapped[str] = mapped_column(String(20), default="user")
    active: Mapped[bool] = mapped_column(Boolean, default=True)


class UserCategory(Base):
    __tablename__ = "user_categories"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    name: Mapped[str] = mapped_column(String(200))
    normalized_name: Mapped[str] = mapped_column(String(200), index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class Report(Base):
    __tablename__ = "reports"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    month: Mapped[int] = mapped_column(Integer)
    year: Mapped[int] = mapped_column(Integer)
    person: Mapped[str] = mapped_column(String(160))
    filename: Mapped[str] = mapped_column(String(255))
    path: Mapped[str] = mapped_column(String(500))
    status: Mapped[str] = mapped_column(String(20), default="active", index=True)
    activity_count: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class ReportActivity(Base):
    __tablename__ = "report_activities"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    report_id: Mapped[int] = mapped_column(ForeignKey("reports.id"), index=True)
    actividad: Mapped[str] = mapped_column(String(500))
    cliente: Mapped[str] = mapped_column(String(300), default="")
    fecha_desde: Mapped[str] = mapped_column(String(40), default="")
    fecha_hasta: Mapped[str] = mapped_column(String(40), default="")
    descripcion: Mapped[str] = mapped_column(Text, default="")
    fingerprint: Mapped[str] = mapped_column(String(64), index=True)


class ReportDraft(Base):
    __tablename__ = "report_drafts"

    id: Mapped[str] = mapped_column(String(80), primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    month: Mapped[int] = mapped_column(Integer)
    year: Mapped[int] = mapped_column(Integer)
    person: Mapped[str] = mapped_column(String(160))
    rows_json: Mapped[str] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class RegisterRequest(BaseModel):
    username: str = Field(min_length=3, max_length=80)
    password: str = Field(min_length=8, max_length=128)


class LoginRequest(RegisterRequest):
    pass


class ReportRow(BaseModel):
    actividad: str = ""
    cliente: str = ""
    fecha_desde: str = ""
    fecha_hasta: str = ""
    descripcion: str = ""


class ReportRequest(BaseModel):
    month: int = Field(ge=1, le=12)
    year: int = Field(ge=2000, le=2100)
    person: str = "GENERAL"
    rows: list[ReportRow] = Field(default_factory=list)


class DraftConfirmRequest(BaseModel):
    draft_id: str
    rows: list[ReportRow] = Field(default_factory=list)
    existing_ids: list[int] = Field(default_factory=list)
    strategy: str = "replace"


class CategoryRequest(BaseModel):
    name: str = Field(min_length=2, max_length=200)


class CategorySuggestRequest(BaseModel):
    actividad: str
    descripcion: str = ""


class SuggestionRequest(BaseModel):
    month: int = Field(ge=1, le=12)
    year: int = Field(ge=2000, le=2100)
    person: str = Field(min_length=2, max_length=160)
    category: str = ""


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 180_000)
    return f"pbkdf2_sha256$180000${salt.hex()}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algorithm, rounds, salt_hex, digest_hex = stored.split("$")
        if algorithm != "pbkdf2_sha256":
            return False
        candidate = hashlib.pbkdf2_hmac(
            "sha256", password.encode(), bytes.fromhex(salt_hex), int(rounds)
        )
        return hmac.compare_digest(candidate.hex(), digest_hex)
    except (ValueError, TypeError):
        return False


def create_token(user: User) -> str:
    return jwt.encode(
        {"sub": str(user.id), "username": user.username, "role": user.role, "exp": int(time.time()) + 86_400},
        JWT_SECRET,
        algorithm="HS256",
    )


def current_user(token: str | None, db: Session) -> User:
    if not token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Sesión requerida")
    try:
        raw_token = token.strip()
        if raw_token.lower().startswith("bearer "):
            raw_token = raw_token[7:].strip()
        payload = jwt.decode(raw_token, JWT_SECRET, algorithms=["HS256"])
        user = db.get(User, int(payload["sub"]))
    except (jwt.InvalidTokenError, KeyError, ValueError):
        user = None
    if not user or not user.active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Sesión no válida")
    return user


def normalize_person(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "_", value.upper()).strip("_")
    return value or "SIN_PERSONA"


def normalize_category(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().casefold()


def normalized_filename(filename: str, month: int, year: int) -> tuple[str, str | None]:
    original = Path(filename).name
    extension = Path(original).suffix.lower()
    stem = Path(original).stem.upper().replace("–", "-").replace("_", "-")
    parts = [part for part in stem.split("-") if part]
    person = "SIN_PERSONA"
    warning = None
    if len(parts) >= 6 and parts[0] in {"ALT", "AT"}:
        person = normalize_person("_".join(parts[6:]))
        if parts[0] != "ALT":
            warning = "El prefijo fue normalizado a ALT."
        if len(parts) < 7 or parts[4] != MONTHS[month] or parts[5] != str(year):
            warning = "El nombre del archivo no coincide con el mes/año seleccionado."
    else:
        warning = "El nombre no sigue el formato ALT-INF-ACT-PER-MES-AÑO-NOMBRE_APELLIDO."
    return f"ALT-INF-ACT-PER-{MONTHS[month]}-{year}-{person}{extension}", warning


def ensure_database() -> None:
    for _ in range(20):
        try:
            Base.metadata.create_all(engine)
            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            REPORT_DIR.mkdir(parents=True, exist_ok=True)
            return
        except Exception:
            time.sleep(2)
    raise RuntimeError("No se pudo conectar con PostgreSQL")


def text_from_excel(content: bytes) -> str:
    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            for values in sheet.iter_rows(values_only=True):
                line = " | ".join(str(value).strip() for value in values if value is not None and str(value).strip())
                if line:
                    lines.append(line)
        return "\n".join(lines)
    finally:
        if workbook:
            workbook.close()


def text_from_pdf(content: bytes) -> str:
    document = fitz.open(stream=content, filetype="pdf")
    try:
        return "\n".join(page.get_text() for page in document).strip()
    finally:
        document.close()


def fallback_description(activity: str) -> str:
    lower = activity.lower()
    if "no ha sido posible" in lower or "no fue posible" in lower:
        return f"Se documentó la siguiente incidencia: {activity}. Se recomienda resolver la causa identificada antes de continuar con las pruebas."
    if lower.startswith("revisión") or lower.startswith("revision"):
        return f"Se realizó {activity.lower()}, dejando registrada la actividad y su resultado para el seguimiento del informe."
    return f"Se registró y documentó la actividad: {activity}."


def rows_from_text(text: str, person: str = "") -> list[ReportRow]:
    rows: list[ReportRow] = []
    date_pattern = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b")
    ignored_labels = {
        "actividad", "descripcion", "descripción", "tarea", "responsable", "usuario registro",
        "tipo", "estado", "prioridad", "fecha inicio", "fecha entrega", "fechas programadas",
        "registro", "archivado", "baja", "tarea imprevista",
    }
    person_value = re.sub(r"\s+", " ", person).strip().lower()
    all_dates = date_pattern.findall(text)
    all_periods = {(int(value.split("/")[1]), int(value.split("/")[2])) for value in all_dates if value.count("/") == 2}
    default_dates = all_dates if len(all_periods) == 1 else []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip(" -|\t")
        normalized = line.lower().rstrip(":")
        if len(line) < 8 or normalized in ignored_labels:
            continue
        if normalized == person_value or normalized.startswith("view ref") or normalized.startswith("ingresado por"):
            continue
        if date_pattern.fullmatch(line) or line.startswith(("📅", "✅", "🏁")):
            continue
        dates = date_pattern.findall(line)
        parts = [part.strip() for part in line.split("|")]
        activity = parts[0][:180]
        description = " | ".join(parts[1:])[:1000] if len(parts) > 1 else fallback_description(activity)
        rows.append(
            ReportRow(
                actividad=activity,
                cliente="",
                fecha_desde=dates[0] if dates else (default_dates[0] if default_dates else ""),
                fecha_hasta=dates[-1] if len(dates) > 1 else (dates[0] if dates else (default_dates[-1] if default_dates else "")),
                descripcion=description,
            )
        )
        if len(rows) >= 100:
            break
    return rows


def response_text(payload: dict[str, object]) -> str:
    if isinstance(payload.get("output_text"), str):
        return payload["output_text"]
    chunks: list[str] = []
    for item in payload.get("output", []):
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []):
            if isinstance(content, dict) and isinstance(content.get("text"), str):
                chunks.append(content["text"])
    return "\n".join(chunks)


def rows_from_ai(text: str) -> list[ReportRow]:
    match = re.search(r"\[[\s\S]*\]", text)
    if not match:
        return []
    try:
        values = json.loads(match.group(0))
    except json.JSONDecodeError:
        return []
    rows: list[ReportRow] = []
    for value in values if isinstance(values, list) else []:
        if isinstance(value, dict) and str(value.get("actividad", "")).strip():
            row = ReportRow(**{field: str(value.get(field, "")) for field in ReportRow.model_fields})
            if not row.descripcion.strip():
                row.descripcion = fallback_description(row.actividad)
            rows.append(row)
    return rows[:100]


def analyze_with_opencode(files: list[tuple[str, bytes]], description: str, extracted_text: str, month: int, year: int) -> list[ReportRow]:
    if not OPENCODE_API_KEY:
        return []
    prompt = (
        f"Convierte las fuentes en filas de un informe mensual de actividades. El periodo objetivo es {MONTHS[month]} {year}. "
        "Devuelve SOLO un arreglo JSON, sin markdown, con estas claves exactas: "
        "actividad, cliente, fecha_desde, fecha_hasta, descripcion. "
        "No inventes fechas ni clientes; deja esos campos vacios si no aparecen. "
        "Escribe una descripcion profesional, completa y util para cada actividad; la descripcion NUNCA puede quedar vacia. "
        "Si una fuente contiene campos tecnicos como Responsable, Estado, Prioridad, View Ref o Registro, no los conviertas en actividades. "
        "Une el contexto de una misma tarea en una sola fila y conserva detalles de incidencias, resultados y pendientes.\n\n"
        f"Texto escrito por el usuario:\n{description[:8000]}\n\n"
        f"Texto extraido de PDF o Excel:\n{extracted_text[:12000]}"
    )
    content: list[dict[str, object]] = [{"type": "input_text", "text": prompt}]
    for filename, raw in files:
        extension = Path(filename).suffix.lower()
        encoded = base64.b64encode(raw).decode("ascii")
        if extension in {".jpg", ".jpeg", ".png"}:
            media_type = "image/png" if extension == ".png" else "image/jpeg"
            content.append({"type": "input_image", "image_url": f"data:{media_type};base64,{encoded}"})
        elif extension == ".pdf":
            content.append({"type": "input_file", "filename": filename, "file_data": f"data:application/pdf;base64,{encoded}"})
    request_body = json.dumps({
        "model": OPENCODE_MODEL,
        "input": [{"role": "user", "content": content}],
        "max_output_tokens": 4000,
    }).encode("utf-8")
    request = urllib.request.Request(
        OPENCODE_URL,
        data=request_body,
        headers={
            "Authorization": f"Bearer {OPENCODE_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "ALTURA-NEXO/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return rows_from_ai(response_text(json.loads(response.read().decode("utf-8"))))
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return []


def suggest_category_with_opencode(actividad: str, descripcion: str, existing: list[str]) -> str:
    if not OPENCODE_API_KEY:
        return ""
    prompt = (
        "Sugiere una categoria breve y profesional (Cliente / Proyecto / Servicio) para una actividad "
        "de un informe mensual de actividades. Responde SOLO con el nombre de la categoria, sin comillas, "
        "sin explicaciones y sin texto adicional, maximo 5 palabras. Si alguna de las categorias existentes "
        "encaja, reutilizala exactamente igual.\n"
        f"Categorias existentes: {', '.join(existing) or 'ninguna'}\n"
        f"Actividad: {actividad[:500]}\n"
        f"Descripcion: {descripcion[:1500]}"
    )
    content: list[dict[str, object]] = [{"type": "input_text", "text": prompt}]
    request_body = json.dumps({
        "model": OPENCODE_MODEL,
        "input": [{"role": "user", "content": content}],
        "max_output_tokens": 80,
    }).encode("utf-8")
    request = urllib.request.Request(
        OPENCODE_URL,
        data=request_body,
        headers={
            "Authorization": f"Bearer {OPENCODE_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "ALTURA-NEXO/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            text = response_text(json.loads(response.read().decode("utf-8"))).strip()
            return re.sub(r"\s+", " ", text.strip("\"'").strip())[:200]
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return ""


def parse_date(value: str) -> date | None:
    value = value.strip()
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue
    return None


def period_dates(month: int, year: int) -> tuple[date, date]:
    first = date(year, month, 1)
    next_month = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
    return first, next_month - timedelta(days=1)


def complete_missing_end_dates(rows: list[ReportRow]) -> list[ReportRow]:
    for row in rows:
        if row.fecha_desde.strip() and not row.fecha_hasta.strip():
            row.fecha_hasta = row.fecha_desde
    return rows


def filter_rows_by_period(rows: list[ReportRow], month: int, year: int) -> tuple[list[ReportRow], int, int]:
    period_start, period_end = period_dates(month, year)
    accepted: list[ReportRow] = []
    discarded = 0
    undated = 0
    for row in rows:
        start = parse_date(row.fecha_desde)
        end = parse_date(row.fecha_hasta) or start
        if start and end:
            if end < period_start or start > period_end:
                discarded += 1
                continue
        elif start:
            if start.month != month or start.year != year:
                discarded += 1
                continue
        else:
            undated += 1
        accepted.append(row)
    return accepted, discarded, undated


def business_days(start_value: str, end_value: str) -> int:
    start = parse_date(start_value)
    end = parse_date(end_value) or start
    if not start or not end or end < start:
        return 0
    total = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            total += 1
        current += timedelta(days=1)
    return total


def activity_fingerprint(row: ReportRow) -> str:
    value = "|".join(re.sub(r"\s+", " ", getattr(row, field, "")).strip().lower() for field in ("actividad", "cliente", "fecha_desde", "fecha_hasta"))
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def row_from_activity(activity: ReportActivity) -> ReportRow:
    return ReportRow(
        actividad=activity.actividad,
        cliente=activity.cliente,
        fecha_desde=activity.fecha_desde,
        fecha_hasta=activity.fecha_hasta,
        descripcion=activity.descripcion,
    )


def create_report_file(month: int, year: int, person: str, rows: list[ReportRow]) -> tuple[Path, str]:
    normalized_person = normalize_person(person)
    filename = f"ALT-INF-RES-PER-{MONTHS[month]}-{year}-{normalized_person}.xlsx"
    destination = REPORT_DIR / str(year) / f"{month:02d}" / filename
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    headers = ["No.", "Actividad", "Cliente/Proyecto/Servicio", "Fecha desde", "Fecha hasta", "Descripción"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B87")
    for index, row in enumerate(rows, start=1):
        sheet.append([index, row.actividad, row.cliente, row.fecha_desde, row.fecha_hasta, row.descripcion])
    for index, width in enumerate([8, 30, 30, 16, 16, 60], start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(destination)
    return destination, filename


app = FastAPI(title="ALTURA NEXO API", version="0.1.0")
origins = [item.strip() for item in os.getenv("CORS_ORIGINS", "http://localhost:9000").split(",")]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    ensure_database()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "service": "altura-nexo-api"}


@app.get("/api/bootstrap-status")
def bootstrap_status(db: Session = Depends(get_db)) -> dict[str, bool]:
    return {"needs_registration": db.scalar(select(User.id).limit(1)) is None}


@app.get("/api/users")
def users(db: Session = Depends(get_db)) -> list[dict[str, str | int]]:
    return [
        {"id": user.id, "username": user.username, "role": user.role}
        for user in db.scalars(select(User).where(User.active.is_(True)).order_by(User.username)).all()
    ]


@app.post("/api/register")
def register(payload: RegisterRequest, db: Session = Depends(get_db)) -> dict[str, str]:
    username = payload.username.strip().lower()
    if db.scalar(select(User.id).where(User.username == username)) is not None:
        raise HTTPException(status_code=409, detail="Ese usuario ya existe")
    role = "admin" if db.scalar(select(User.id).limit(1)) is None else "user"
    user = User(username=username, password_hash=hash_password(payload.password), role=role)
    db.add(user)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Ese usuario ya existe")
    return {"message": "Usuario creado", "role": role}


@app.post("/api/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> dict[str, object]:
    user = db.scalar(select(User).where(User.username == payload.username.strip().lower()))
    if not user or not user.active or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    return {"access_token": create_token(user), "user": {"id": user.id, "username": user.username, "role": user.role}}


@app.get("/api/categories")
def categories(
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> list[dict[str, object]]:
    user = current_user(authorization, db)
    values = db.scalars(select(UserCategory).where(UserCategory.user_id == user.id).order_by(UserCategory.name)).all()
    return [{"id": item.id, "name": item.name} for item in values]


@app.post("/api/categories")
def create_category(
    payload: CategoryRequest,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> dict[str, object]:
    user = current_user(authorization, db)
    name = re.sub(r"\s+", " ", payload.name).strip()
    normalized = normalize_category(name)
    existing = db.scalar(select(UserCategory).where(UserCategory.user_id == user.id, UserCategory.normalized_name == normalized))
    if existing:
        return {"id": existing.id, "name": existing.name}
    item = UserCategory(user_id=user.id, name=name, normalized_name=normalized)
    db.add(item)
    db.commit()
    return {"id": item.id, "name": item.name}


@app.post("/api/categories/suggest")
def suggest_category(
    payload: CategorySuggestRequest,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> dict[str, object]:
    user = current_user(authorization, db)
    existing = [
        item.name for item in db.scalars(select(UserCategory).where(UserCategory.user_id == user.id).order_by(UserCategory.name)).all()
    ]
    name = suggest_category_with_opencode(payload.actividad, payload.descripcion, existing)
    if not name:
        raise HTTPException(status_code=422, detail="No se pudo sugerir una categoría automáticamente; escríbela manualmente.")
    return {"name": name}


@app.get("/api/people")
def people(
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> list[str]:
    user = current_user(authorization, db)
    values = db.scalars(select(Report.person).where(Report.user_id == user.id, Report.status == "active").distinct().order_by(Report.person)).all()
    return list(values)


@app.post("/api/suggestions")
def suggestions(
    payload: SuggestionRequest,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> dict[str, object]:
    user = current_user(authorization, db)
    categories_for_user = db.scalars(select(UserCategory).where(UserCategory.user_id == user.id).order_by(UserCategory.name)).all()
    reports = db.scalars(select(Report).where(Report.user_id == user.id, Report.status == "active").order_by(Report.created_at.desc())).all()
    previous: list[str] = []
    for report in reports[:12]:
        previous.extend(activity.actividad for activity in db.scalars(select(ReportActivity).where(ReportActivity.report_id == report.id)).all())
    category = payload.category.strip()
    context = "\n".join(f"- {item}" for item in previous[:60]) or "No hay actividades anteriores registradas."
    category_context = ", ".join(item.name for item in categories_for_user) or "Sin categorías guardadas"
    prompt = (
        f"Sugiere actividades profesionales para un informe de {MONTHS[payload.month]} {payload.year} de la persona {payload.person}. "
        f"Proyecto o categoría seleccionada: {category or 'general'}. Categorías existentes: {category_context}.\n"
        "Estas son sugerencias basadas en el historial, no hechos confirmados. Devuelve SOLO JSON con claves "
        "actividad, cliente, fecha_desde, fecha_hasta, descripcion. No inventes fechas ni afirmes que algo ocurrió. "
        "Genera como máximo 8 actividades concretas que el usuario pueda revisar y aceptar. "
        "La descripción debe explicar qué debería documentarse en esa actividad.\n"
        f"Historial previo:\n{context}"
    )
    rows = analyze_with_opencode([], prompt, "", payload.month, payload.year)
    if not rows:
        if not category and not categories_for_user:
            raise HTTPException(status_code=422, detail="Guarda primero una categoría o carga una fuente para generar sugerencias.")
        fallback_categories = [category] if category else [item.name for item in categories_for_user[:4]]
        rows = [ReportRow(
            actividad=f"Seguimiento de {item}",
            cliente=item,
            descripcion=f"Sugerencia para documentar avances, incidencias y próximos pasos relacionados con {item}.",
        ) for item in fallback_categories]
    existing = db.scalar(select(Report).where(
        Report.user_id == user.id,
        Report.month == payload.month,
        Report.year == payload.year,
        Report.person == payload.person.strip(),
        Report.status == "active",
    ))
    existing_activities = []
    if existing:
        existing_activities = [{
            "id": activity.id,
            "actividad": activity.actividad,
            "cliente": activity.cliente,
            "fecha_desde": activity.fecha_desde,
            "fecha_hasta": activity.fecha_hasta,
            "descripcion": activity.descripcion,
        } for activity in db.scalars(select(ReportActivity).where(ReportActivity.report_id == existing.id).order_by(ReportActivity.id)).all()]
    draft_id = secrets.token_urlsafe(24)
    db.add(ReportDraft(
        id=draft_id,
        user_id=user.id,
        month=payload.month,
        year=payload.year,
        person=payload.person.strip(),
        rows_json=json.dumps([row.model_dump() for row in rows], ensure_ascii=False),
    ))
    db.commit()
    return {
        "draft_id": draft_id,
        "rows": [row.model_dump() for row in rows],
        "suggestion": True,
        "existing_report": {"id": existing.id, "activity_count": existing.activity_count, "created_at": existing.created_at.isoformat()} if existing else None,
        "existing_activities": existing_activities,
    }


@app.post("/api/uploads")
async def upload_files(
    month: Annotated[int, Form(ge=1, le=12)],
    year: Annotated[int, Form(ge=2000, le=2100)],
    files: list[UploadFile] = File(...),
    authorization: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    current_user(authorization, db)
    current_year = datetime.now().year
    warnings: list[str] = []
    saved: list[dict[str, str]] = []
    target = UPLOAD_DIR / str(year) / f"{month:02d}"
    target.mkdir(parents=True, exist_ok=True)
    for index, upload in enumerate(files, start=1):
        extension = Path(upload.filename or "").suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            warnings.append(f"{upload.filename}: tipo de archivo no permitido.")
            continue
        name, warning = normalized_filename(upload.filename or f"documento-{index}", month, year)
        destination = target / f"{index:03d}-{name}"
        destination.write_bytes(await upload.read())
        if warning:
            warnings.append(f"{upload.filename}: {warning}")
        saved.append({"original": upload.filename or "", "stored": str(destination.relative_to(DATA_DIR))})
    if year != current_year:
        warnings.append(f"El año seleccionado es {year}; el año actual es {current_year}.")
    return {"month": MONTHS[month], "year": year, "saved": saved, "warnings": warnings}


@app.post("/api/reports/analyze")
async def analyze_report_sources(
    month: Annotated[int, Form(ge=1, le=12)],
    year: Annotated[int, Form(ge=2000, le=2100)],
    person: Annotated[str, Form()],
    description: Annotated[str, Form()] = "",
    files: list[UploadFile] | None = File(default=None),
    authorization: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
):
    user = current_user(authorization, db)
    source_files = files or []
    if not source_files and not description.strip():
        raise HTTPException(status_code=422, detail="Agrega una foto, PDF, Excel o escribe las actividades.")

    warnings: list[str] = []
    raw_sources: list[tuple[str, bytes]] = []
    extracted_parts: list[str] = []
    target = UPLOAD_DIR / str(year) / f"{month:02d}"
    target.mkdir(parents=True, exist_ok=True)
    for index, upload in enumerate(source_files, start=1):
        filename = upload.filename or f"fuente-{index}"
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            warnings.append(f"{filename}: tipo de archivo no permitido.")
            continue
        content = await upload.read()
        raw_sources.append((filename, content))
        stored_name, warning = normalized_filename(filename, month, year)
        (target / f"{index:03d}-{stored_name}").write_bytes(content)
        if warning:
            warnings.append(f"{filename}: {warning}")
        try:
            if extension in {".xlsx", ".xls"}:
                extracted = text_from_excel(content)
            elif extension == ".pdf":
                extracted = text_from_pdf(content)
            else:
                extracted = ""
            if extracted:
                extracted_parts.append(f"[{filename}]\n{extracted}")
        except Exception:
            warnings.append(f"{filename}: no se pudo extraer texto local.")

    extracted_text = "\n\n".join(extracted_parts)
    rows = analyze_with_opencode(raw_sources, description, extracted_text, month, year)
    if not rows:
        rows = rows_from_text(description, person) + rows_from_text(extracted_text, person)
        if raw_sources and not OPENCODE_API_KEY:
            warnings.append("OpenCode no esta configurado; las imagenes no se pudieron leer automaticamente.")
        elif OPENCODE_API_KEY and raw_sources:
            warnings.append("No se pudo usar el modelo; se genero el informe con la lectura local disponible.")
    rows, discarded, undated = filter_rows_by_period(complete_missing_end_dates(rows), month, year)
    if discarded:
        warnings.append(f"Se descartaron {discarded} actividad(es) fuera de {MONTHS[month]} {year}.")
    if undated:
        warnings.append(f"{undated} actividad(es) no tiene(n) fecha confirmada; revísala(s) antes de aceptar.")
    if not rows:
        raise HTTPException(
            status_code=422,
            detail="No se encontraron actividades. Escribe una descripcion o configura OPENCODE_API_KEY para leer imagenes.",
        )

    draft_id = secrets.token_urlsafe(24)
    existing = db.scalar(select(Report).where(
        Report.user_id == user.id,
        Report.month == month,
        Report.year == year,
        Report.person == person.strip(),
        Report.status == "active",
    ))
    existing_activities = []
    if existing:
        existing_activities = [{
            "id": activity.id,
            "actividad": activity.actividad,
            "cliente": activity.cliente,
            "fecha_desde": activity.fecha_desde,
            "fecha_hasta": activity.fecha_hasta or activity.fecha_desde,
            "descripcion": activity.descripcion,
        } for activity in db.scalars(select(ReportActivity).where(ReportActivity.report_id == existing.id).order_by(ReportActivity.id)).all()]
    db.add(ReportDraft(
        id=draft_id,
        user_id=user.id,
        month=month,
        year=year,
        person=person.strip(),
        rows_json=json.dumps([row.model_dump() for row in rows], ensure_ascii=False),
    ))
    db.commit()
    return {
        "draft_id": draft_id,
        "month": MONTHS[month],
        "year": year,
        "person": person.strip(),
        "rows": [row.model_dump() for row in rows],
        "warnings": warnings,
        "existing_report": {
            "id": existing.id,
            "activity_count": existing.activity_count,
            "created_at": existing.created_at.isoformat(),
        } if existing else None,
        "existing_activities": existing_activities,
    }


@app.post("/api/reports/confirm")
def confirm_report(
    payload: DraftConfirmRequest,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
):
    user = current_user(authorization, db)
    draft = db.get(ReportDraft, payload.draft_id)
    if not draft or draft.user_id != user.id:
        raise HTTPException(status_code=404, detail="Borrador no encontrado")
    if payload.strategy not in {"replace", "merge"}:
        raise HTTPException(status_code=400, detail="Estrategia no válida")
    rows, _, _ = filter_rows_by_period(complete_missing_end_dates(payload.rows), draft.month, draft.year)

    existing = db.scalar(select(Report).where(
        Report.user_id == user.id,
        Report.month == draft.month,
        Report.year == draft.year,
        Report.person == draft.person,
        Report.status == "active",
    ))
    if existing and payload.strategy == "merge":
        old_rows = [row_from_activity(item) for item in db.scalars(select(ReportActivity).where(ReportActivity.report_id == existing.id, ReportActivity.id.in_(payload.existing_ids))).all()]
        combined: dict[str, ReportRow] = {activity_fingerprint(row): row for row in old_rows}
        combined.update({activity_fingerprint(row): row for row in rows})
        rows = list(combined.values())
    if not rows:
        raise HTTPException(status_code=422, detail="Selecciona al menos una actividad válida")
    if existing:
        existing.status = "replaced"

    known_categories = {
        item.normalized_name for item in db.scalars(select(UserCategory).where(UserCategory.user_id == user.id)).all()
    }
    for row in rows:
        category = re.sub(r"\s+", " ", row.cliente).strip()
        normalized_category = normalize_category(category)
        if category and normalized_category not in known_categories:
            db.add(UserCategory(user_id=user.id, name=category, normalized_name=normalized_category))
            known_categories.add(normalized_category)

    destination, filename = create_report_file(draft.month, draft.year, draft.person, rows)
    report = Report(
        user_id=user.id,
        month=draft.month,
        year=draft.year,
        person=draft.person,
        filename=filename,
        path=str(destination.relative_to(DATA_DIR)),
        activity_count=len(rows),
        status="active",
    )
    db.add(report)
    db.flush()
    for row in rows:
        db.add(ReportActivity(
            report_id=report.id,
            actividad=row.actividad,
            cliente=row.cliente,
            fecha_desde=row.fecha_desde,
            fecha_hasta=row.fecha_hasta,
            descripcion=row.descripcion,
            fingerprint=activity_fingerprint(row),
        ))
    db.delete(draft)
    db.commit()
    return FileResponse(
        destination,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"X-Report-Rows": str(len(rows))},
    )


@app.get("/api/reports")
def report_history(
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> list[dict[str, object]]:
    user = current_user(authorization, db)
    reports = db.scalars(select(Report).where(Report.user_id == user.id, Report.status == "active").order_by(Report.created_at.desc())).all()
    return [{
        "id": report.id,
        "month": MONTHS[report.month],
        "month_number": report.month,
        "year": report.year,
        "person": report.person,
        "activity_count": report.activity_count,
        "created_at": report.created_at.isoformat(),
    } for report in reports]


@app.get("/api/reports/{report_id}/download")
def download_report(
    report_id: int,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
):
    user = current_user(authorization, db)
    report = db.scalar(select(Report).where(Report.id == report_id, Report.user_id == user.id, Report.status == "active"))
    if not report:
        raise HTTPException(status_code=404, detail="Informe no encontrado")
    destination = DATA_DIR / report.path
    if not destination.exists():
        raise HTTPException(status_code=404, detail="Archivo del informe no encontrado")
    return FileResponse(destination, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=report.filename)


@app.get("/api/dashboard")
def dashboard(
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
) -> dict[str, object]:
    user = current_user(authorization, db)
    reports = db.scalars(select(Report).where(Report.user_id == user.id, Report.status == "active").order_by(Report.created_at)).all()
    report_lookup = {report.id: report for report in reports}
    activities = []
    for report in reports:
        activities.extend(db.scalars(select(ReportActivity).where(ReportActivity.report_id == report.id)).all())
    categories: dict[str, dict[str, int | str]] = {}
    people: dict[str, int] = {}
    monthly: dict[tuple[int, int], dict[str, object]] = {}
    daily: dict[str, dict[str, object]] = {}
    timeline: list[dict[str, object]] = []
    for activity in activities:
        category = activity.cliente.strip() or "Sin categoría"
        duration = business_days(activity.fecha_desde, activity.fecha_hasta)
        item = categories.setdefault(category, {"category": category, "activity_count": 0, "business_days": 0})
        item["activity_count"] = int(item["activity_count"]) + 1
        item["business_days"] = int(item["business_days"]) + duration
        report = report_lookup.get(activity.report_id)
        if report:
            people[report.person] = people.get(report.person, 0) + 1
            period_key = (report.year, report.month)
            period = monthly.setdefault(period_key, {
                "year": report.year,
                "month": report.month,
                "period": f"{MONTHS[report.month]} {report.year}",
                "activity_count": 0,
                "business_days": 0,
                "projects": {},
            })
            period["activity_count"] = int(period["activity_count"]) + 1
            period["business_days"] = int(period["business_days"]) + duration
            projects = period["projects"]
            if isinstance(projects, dict):
                project = projects.setdefault(category, {"activity_count": 0, "business_days": 0})
                project["activity_count"] = int(project["activity_count"]) + 1
                project["business_days"] = int(project["business_days"]) + duration
        start = parse_date(activity.fecha_desde)
        end = parse_date(activity.fecha_hasta) or start
        if start and end and end >= start:
            current = start
            while current <= end:
                if current.weekday() < 5:
                    day_key = current.isoformat()
                    day = daily.setdefault(day_key, {
                        "date": day_key,
                        "label": current.strftime("%d/%m"),
                        "activity_count": 0,
                        "projects": {},
                    })
                    day["activity_count"] = int(day["activity_count"]) + 1
                    day_projects = day["projects"]
                    if isinstance(day_projects, dict):
                        day_project = day_projects.setdefault(category, 0)
                        day_projects[category] = int(day_project) + 1
                current += timedelta(days=1)
        timeline.append({
            "actividad": activity.actividad,
            "categoria": category,
            "fecha_desde": activity.fecha_desde,
            "fecha_hasta": activity.fecha_hasta,
            "business_days": duration,
            "descripcion": activity.descripcion,
        })
    for report in reports:
        first_day, last_day = period_dates(report.month, report.year)
        current_day = first_day
        while current_day <= last_day:
            day_key = current_day.isoformat()
            daily.setdefault(day_key, {
                "date": day_key,
                "label": current_day.strftime("%d/%m"),
                "activity_count": 0,
                "projects": {},
            })
            current_day += timedelta(days=1)
    timeline.sort(key=lambda item: parse_date(str(item["fecha_desde"])) or date.max)
    project_comparison = [{
        "project": item["category"],
        "activity_count": item["activity_count"],
        "business_days": item["business_days"],
    } for item in sorted(categories.values(), key=lambda item: (-int(item["activity_count"]), -int(item["business_days"])))]
    return {
        "report_count": len(reports),
        "activity_count": len(activities),
        "category_count": len(categories),
        "business_days": sum(int(item["business_days"]) for item in categories.values()),
        "categories": sorted(categories.values(), key=lambda item: (-int(item["business_days"]), -int(item["activity_count"]))),
        "projects": project_comparison,
        "monthly_series": [monthly[key] for key in sorted(monthly)],
        "daily_series": [daily[key] for key in sorted(daily)],
        "people": [{"person": person, "activity_count": count} for person, count in sorted(people.items())],
        "timeline": timeline,
    }


@app.post("/api/reports/generate")
def generate_report(
    payload: ReportRequest,
    db: Session = Depends(get_db),
    authorization: Annotated[str | None, Header()] = None,
):
    current_user(authorization, db)
    destination, filename = create_report_file(payload.month, payload.year, payload.person, payload.rows)
    return FileResponse(
        destination,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
